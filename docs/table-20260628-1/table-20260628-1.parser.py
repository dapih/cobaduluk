"""
Generated parser for job table-20260628-1 (3-level revision).
Hierarchy: Entri → RiskGroup (tingkat_risiko) → KewenanganGroup (kewenangan).
Every populated source row feeds the current entry's current risk group's
current kewenangan group.  No row is dropped.

Production parsers use ${CLAUDE_PLUGIN_ROOT}; here docs/ sits under the plugin
root so we resolve scripts/ two levels up from this file.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, os.path.join(ROOT, "scripts"))

import re
import openpyxl
from parser_lib import clean, dehyphenate, nest_by_pattern, dedupe, as_int_str, write_json

_LEADING_DASH = re.compile(r"^-\s+")

JOB = "table-20260628-1"
SRC = os.path.join(HERE, JOB + ".xlsx")
OUT = os.path.join(HERE, JOB + ".json")

LEVELS = [("num", r"^\d+\."), ("alpha", r"^[a-z]\."), ("paren", r"^\d+\)")]
PLACEHOLDERS = ["-", "–", "—"]

# Domain-specific authority normalization.
# "Menteri/..." → "Menteri/Kepala Badan"; "Bupati/..." → "Bupati/Walikota".
# "Kepala Badan" is excluded from _TRIGGER_KEW because it appears as the trailing
# fragment of a "Menteri/\nKepala Badan" split across two rows, not a standalone
# authority — triggering on it would open a spurious third kewenangan group.
_TRIGGER_KEW = {
    "Bupati", "Walikota", "Bupati/Walikota",
    "Gubernur", "Menteri", "Menteri/Kepala Badan", "Presiden",
}


def norm_kew(v):
    """Normalize a raw kewenangan cell to its canonical form, or None."""
    v = clean(v)
    if not v:
        return None
    if v.startswith("Bupati/"):
        return "Bupati/Walikota"
    if v.startswith("Menteri/"):
        return "Menteri/Kepala Badan"
    return v


def norm(v):
    """Clean then dehyphenate with reduplication protection."""
    return dehyphenate(clean(v), merge_no_space=True)


def norm_skala(v):
    """Like norm(), but also strips leading '- ' list markers common in scale cells."""
    v = norm(v)
    if not v:
        return None
    v = _LEADING_DASH.sub("", v).strip()
    return v or None


def add_flat(lst, v):
    v = norm(v)
    if v and v not in PLACEHOLDERS:
        lst.append(v)


def new_risk_group(risiko, perizinan):
    return {
        "nilai": norm(risiko),
        "perizinan_berusaha": norm(perizinan),
        "_kew": {},   # key=canonical kewenangan → KewenanganGroup; converted to list in finalize
    }


def new_kew_group(kw_val, skala_raw=None, jangka_raw=None):
    s = norm_skala(skala_raw)
    return {
        "nilai": kw_val,
        "skala_usaha": [s] if s and s not in PLACEHOLDERS else [],
        "jangka_waktu_penerbitan": norm(jangka_raw),
        "_persyaratan": [],
        "_kewajiban": [],
        "parameter": [],
        "pb_umku": [],
    }


def open_or_switch_kew(kw, skala_raw, jangka_raw, kew_dict):
    """Return the KewenanganGroup for kw, creating it if needed."""
    if kw in kew_dict:
        obj = kew_dict[kw]
        s = norm_skala(skala_raw)
        if s and s not in PLACEHOLDERS and s not in obj["skala_usaha"]:
            obj["skala_usaha"].append(s)
        if jangka_raw and not obj["jangka_waktu_penerbitan"]:
            obj["jangka_waktu_penerbitan"] = norm(jangka_raw)
    else:
        obj = new_kew_group(kw, skala_raw, jangka_raw)
        kew_dict[kw] = obj
    return obj


def finalize(entry):
    for rg in entry["tingkat_risiko"]:
        kew_list = []
        for kw_obj in rg.pop("_kew").values():
            kw_obj["persyaratan"] = nest_by_pattern(
                kw_obj.pop("_persyaratan"), LEVELS, drop=PLACEHOLDERS, orphan="attach"
            )
            kw_obj["kewajiban"] = nest_by_pattern(
                kw_obj.pop("_kewajiban"), LEVELS, drop=PLACEHOLDERS, orphan="attach"
            )
            kw_obj["parameter"] = dedupe(kw_obj["parameter"])
            kw_obj["pb_umku"] = dedupe(kw_obj["pb_umku"])
            kew_list.append(kw_obj)
        rg["kewenangan"] = kew_list
    return entry


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["KKP1"]

    row1 = [clean(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    title = next((c for c in row1 if c), None)

    entries = []
    cur = None
    cur_rg = None
    cur_kew = None
    rows_seen = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        rows_seen += 1
        A, B, C, D, E, F, G, H, I, J, K, L, M = (list(row) + [None] * 13)[:13]

        kw = norm_kew(clean(M))

        if isinstance(A, int) and not isinstance(A, bool):
            # ── Entry boundary ───────────────────────────────────────────────
            if cur is not None:
                entries.append(finalize(cur))
            cur = {
                "no": A,
                "kode_kbli": as_int_str(B),
                "judul_kbli": norm(C),
                "ruang_lingkup": norm(D),
                "tingkat_risiko": [],
            }
            cur_rg = new_risk_group(F, G)
            cur["tingkat_risiko"].append(cur_rg)
            cur_kew = None
            if kw:
                cur_kew = open_or_switch_kew(kw, E, I, cur_rg["_kew"])

        else:
            f = norm(F)
            if f:
                # ── F is populated (risk-group boundary OR same-value re-entry) ─
                # Matches original: any F-present row calls _switch_l3 WITH jangka.
                if cur_rg is None or f != cur_rg["nilai"]:
                    cur_rg = new_risk_group(F, G)
                    cur["tingkat_risiko"].append(cur_rg)
                    cur_kew = None
                if kw:
                    cur_kew = open_or_switch_kew(kw, E, I, cur_rg["_kew"])
                # If kw is None, skala is not added here (matches original's
                # _switch_l3 early-return when kw is absent).
            else:
                # ── Pure continuation row (F empty) ──────────────────────────
                if kw and kw in _TRIGGER_KEW and (cur_kew is None or kw != cur_kew["nilai"]):
                    # Only switch kew for known trigger authorities;
                    # "Kepala Badan" alone is a split-row fragment, not a new authority.
                    # Jangka not passed: I on continuation rows is unreliable.
                    cur_kew = open_or_switch_kew(kw, E, None, cur_rg["_kew"])
                elif cur_kew is not None and E is not None:
                    # No switch: accumulate skala into the current kew group
                    s = norm_skala(E)
                    if s and s not in PLACEHOLDERS and s not in cur_kew["skala_usaha"]:
                        cur_kew["skala_usaha"].append(s)

        # Accumulate detail arrays into the current kewenangan group
        if cur_kew is not None:
            add_flat(cur_kew["_persyaratan"], H)
            add_flat(cur_kew["_kewajiban"], J)
            k = norm(K)
            if k and k not in PLACEHOLDERS:
                k = k.lstrip("- ").strip()
                if k and k not in PLACEHOLDERS:
                    cur_kew["pb_umku"].append(k)
            add_flat(cur_kew["parameter"], L)

    if cur is not None:
        entries.append(finalize(cur))

    doc = {"entri": entries}
    if title:
        doc = {"judul_lampiran": title, "entri": entries}

    write_json(doc, OUT)

    total_rg = sum(len(e["tingkat_risiko"]) for e in entries)
    total_kew = sum(
        len(rg["kewenangan"]) for e in entries for rg in e["tingkat_risiko"]
    )
    print(f"rows_seen={rows_seen} entries={len(entries)} risk_groups={total_rg} kew_groups={total_kew}")


if __name__ == "__main__":
    main()
