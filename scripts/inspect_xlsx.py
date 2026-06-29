"""
inspect_xlsx.py - structural inspector for Excel workbooks.

Emits a COMPACT, token-bounded report of a workbook's structure so an AI can
plan a conversion without ever loading the full sheet into context. Reports:
sheet list with dimensions, merged-cell ranges, a detected header row, per-column
profiles (header text, fill rate, distinct/sample values, inferred type), a small
verbatim data sample, and a blank-gap / multi-table heuristic.

Usage:
    python inspect_xlsx.py <file.xlsx> [--sheet NAME] [--header-row N]
        [--max-cols 60] [--sample-rows 10] [--top-values 8]
        [--scan-rows 5000] [--out PREFIX] [--json] [--md]

With --out PREFIX writes PREFIX.inspect.json and/or PREFIX.inspect.md
(default both). Without --out, prints markdown to stdout. Zero AI tokens.
"""
import argparse
import datetime
import json
import sys
from collections import Counter

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def cell_type(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    if isinstance(v, (datetime.datetime, datetime.date)):
        return "date"
    return "str"


def cell_str(v, limit=60):
    if v is None:
        return ""
    s = " ".join(str(v).replace("\xa0", " ").split())
    return s if len(s) <= limit else s[: limit - 1] + "…"


def is_blank(row):
    return all(v is None or str(v).strip() == "" for v in row)


def detect_header_row(grid, max_probe=15):
    """Pick the densest texty row among the first `max_probe` rows (1-indexed)."""
    best_idx, best_score = 1, -1
    for i, row in enumerate(grid[:max_probe], start=1):
        texty = sum(1 for c in row if isinstance(c, str) and c.strip())
        nonempty = sum(1 for c in row if c is not None and str(c).strip())
        score = texty * 2 + nonempty
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def profile_column(header, values, top_values):
    nonempty = [v for v in values if v is not None and str(v).strip() != ""]
    fill = round(len(nonempty) / len(values), 3) if values else 0.0
    types = Counter(cell_type(v) for v in nonempty)
    inferred = types.most_common(1)[0][0] if types else None

    counter = Counter()
    distinct_capped = False
    for v in nonempty:
        key = cell_str(v, 40)
        if key in counter or len(counter) < 200:
            counter[key] += 1
        else:
            distinct_capped = True
    max_len = max((len(str(v)) for v in nonempty), default=0)
    return {
        "header": cell_str(header, 80),
        "fill_rate": fill,
        "nonempty": len(nonempty),
        "distinct": ("200+" if distinct_capped else len(counter)),
        "inferred_type": inferred,
        "type_mix": dict(types),
        "max_len": max_len,
        "top_values": [{"value": v, "count": c} for v, c in counter.most_common(top_values)],
    }


def inspect_sheet(ws, args):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    ncols = min(max_col, args.max_cols) if max_col else 0

    grid = []
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        grid.append(list(row[:ncols]) if ncols else list(row))
        if r >= args.scan_rows:
            break
    scanned = len(grid)

    header_row = args.header_row or detect_header_row(grid)
    header_idx0 = header_row - 1
    headers = grid[header_idx0] if 0 <= header_idx0 < len(grid) else [None] * ncols
    data = grid[header_row:]

    columns = []
    for c in range(ncols):
        col_vals = [row[c] if c < len(row) else None for row in data]
        prof = profile_column(headers[c] if c < len(headers) else None, col_vals, args.top_values)
        prof["column"] = get_column_letter(c + 1)
        columns.append(prof)

    blank_rows = [header_row + 1 + i for i, row in enumerate(data) if is_blank(row)]

    sample = []
    for i, row in enumerate(data):
        if is_blank(row):
            continue
        sample.append({"row": header_row + 1 + i, "cells": [cell_str(v) for v in row]})
        if len(sample) >= args.sample_rows:
            break

    try:
        merged = [str(rng) for rng in ws.merged_cells.ranges]
    except Exception:
        merged = []

    return {
        "name": ws.title,
        "dimensions": {"max_row": max_row, "max_col": max_col, "scanned_rows": scanned},
        "truncated": max_row > scanned or max_col > ncols,
        "header_row": header_row,
        "merged_cells": {"count": len(merged), "ranges": merged[:50]},
        "blank_row_gaps": {"count": len(blank_rows), "rows": blank_rows[:30]},
        "columns": columns,
        "sample_rows": sample,
    }


def render_md(report):
    L = []
    wb = report["workbook"]
    s = report["sheet"]
    L.append(f"# Excel inspection: {wb['file']}")
    L.append("")
    L.append("- Sheets ({}): {}".format(
        len(wb["sheets"]),
        ", ".join(f"`{x['name']}` ({x['max_row']}x{x['max_col']})" for x in wb["sheets"]),
    ))
    if len(wb["sheets"]) > 1:
        L.append(f"- **Multiple sheets present** - inspected `{s['name']}`; use --sheet to pick another.")
    d = s["dimensions"]
    L.append("")
    L.append(f"## Sheet `{s['name']}`")
    L.append("")
    L.append(f"- Dimensions: {d['max_row']} rows x {d['max_col']} cols (scanned {d['scanned_rows']})"
             + (" **[truncated]**" if s["truncated"] else ""))
    L.append(f"- Detected header row: **{s['header_row']}** (override with --header-row)")
    mc = s["merged_cells"]
    L.append(f"- Merged cell ranges: {mc['count']}"
             + (f" (e.g. {', '.join(mc['ranges'][:6])})" if mc["count"] else ""))
    gaps = s["blank_row_gaps"]
    if gaps["count"]:
        L.append(f"- Blank-row gaps: {gaps['count']} (rows {gaps['rows'][:10]}) "
                 "- possible stacked tables; verify before converting.")
    L.append("")
    L.append("### Columns")
    L.append("")
    L.append("| Col | Header | Type | Fill | Distinct | MaxLen | Sample values |")
    L.append("|---|---|---|---|---|---|---|")
    for c in s["columns"]:
        samples = "; ".join(str(tv["value"]) for tv in c["top_values"][:4])
        L.append(f"| {c['column']} | {c['header'] or '-'} | {c['inferred_type'] or '-'} | "
                 f"{c['fill_rate']} | {c['distinct']} | {c['max_len']} | {samples} |")
    L.append("")
    L.append(f"### Sample rows (first {len(s['sample_rows'])} non-empty data rows)")
    L.append("")
    for row in s["sample_rows"]:
        L.append(f"- r{row['row']}: " + " | ".join(row["cells"]))
    L.append("")
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser(description="Inspect the structure of an Excel workbook.")
    ap.add_argument("file")
    ap.add_argument("--sheet")
    ap.add_argument("--header-row", type=int, default=0)
    ap.add_argument("--max-cols", type=int, default=60)
    ap.add_argument("--sample-rows", type=int, default=10)
    ap.add_argument("--top-values", type=int, default=8)
    ap.add_argument("--scan-rows", type=int, default=5000)
    ap.add_argument("--out")
    ap.add_argument("--json", action="store_true", dest="as_json")
    ap.add_argument("--md", action="store_true")
    args = ap.parse_args()

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    try:
        wb = openpyxl.load_workbook(args.file, data_only=True)
    except Exception as e:
        print(f"ERROR: cannot open workbook: {e}", file=sys.stderr)
        return 2

    sheets_meta = [
        {"name": ws.title, "max_row": ws.max_row or 0, "max_col": ws.max_column or 0}
        for ws in wb.worksheets
    ]
    target = args.sheet or wb.active.title
    if target not in wb.sheetnames:
        print(f"ERROR: sheet '{target}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        return 2

    report = {
        "workbook": {"file": args.file, "sheets": sheets_meta},
        "sheet": inspect_sheet(wb[target], args),
    }

    md = render_md(report)
    want_json = args.as_json or not args.md
    want_md = args.md or not args.as_json

    if args.out:
        if want_json:
            with open(args.out + ".inspect.json", "w", encoding="utf-8") as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
        if want_md:
            with open(args.out + ".inspect.md", "w", encoding="utf-8") as f:
                f.write(md)
        print(f"Wrote {args.out}.inspect.json / .inspect.md")
    elif args.as_json and not args.md:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(md)
    return 0


if __name__ == "__main__":
    sys.exit(main())
